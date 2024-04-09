import io
import os
import re
import base64
import ctypes
import locale
import random
import datetime
import openpyxl
import threading
import subprocess
import webbrowser
import configparser
import tkinter as tk
from tkinter.ttk import *
from PIL import Image, ImageTk
from openpyxl.styles import Font, Alignment
from tkinter import ttk, filedialog, messagebox

# 全局变量
UA = ""
PROXY = ""
API_KEY = ""
RANDOM_UA = ""
POC_FOLDER = os.path.join(os.getcwd(), "poc")
NUCLEI_PATH = "nuclei.exe"
YAML_FILES = ""
WORK_FOLDER = os.path.join(os.getcwd(), "work")
ERROR_MESSAGE = ""
MANUALLY_STOP = False
URL_FILE_PATH = ""
SELECT_ALL_POC = False
SELECTED_FILES = []
TREEVIEW_SELECT = False
CONFIG_FILE_PATH = "./work/config.ini"
RANDOM_UA_STATUS = ""
CLOUD_UPLOAD_STATUS = False
CURRENT_LOAD_STATUS = False
CURRENT_SCAN_STATUS = False
CURRENT_SCAN_PROCESS = ""
CURRENT_POCLIST_RESULT = False
CURRENT_VALIDATE_RESULT = False
CURRENT_VALIDATE_STATUS = False
CURRENT_VALIDATE_PROCESS = ""

# 创建主窗口
root = tk.Tk()
# base64编码的图像文件
base64_icon = """
AAABAAEAICAAAAEAIACoEAAAFgAAACgAAAAgAAAAQAAAAAEAIAAAAAAAABAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA29vbAP///wBjY2MAfHx8AExMTABCQkIAZ2dnADo6OgDg4OAAf39/AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA1NTUANbW1gAMDAwAiIiIAF9fXwBwcHAAsLCwAImJiQD///8AAAAAAGtrawC/v78AAAAAAEtLSwD///8BR0dHGTs7OyKAgIAGQkJCAEJCQgChoaEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD///8AAAAAACoqKgC+vr4FVlZWFXBwcA8AAAAAOjo6AAAAAACLi4sAq6urAAAAAAATExMAYGBgFSsrK3cjIyO2IiIivS0tLXmHh4cIb29vACwsLAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHh4eABBQUEAWFhYFikpKYoeHh62ISEhryoqKnRLS0sZGRkZAL6+vgAAAAAAICAgAEtLSyIhISGmISEhpEBAQD1JSUk0JSUlrjU1NT8mJiYAioqKAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAsbGxAAAAAAAvLy9tGxsbqzU1NTUpKSlLHR0dphwcHLM8PDw4AAAAADMzMwBMTEwbHh4eqyMjI5JTU1MSKSkpAAAAAAArKyt5ICAgdwAAAADY2NgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACEhIQA8PDwAyIiIpksLCxfDQ0NAA8PDwB8fHwOKysrdCIiIrpFRUU1nJycCSMjI5ciIiKddXV1ETw8PAAAAAAAAAAAADQ0NF8lJSWT////ApOTkwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAHt7ewC6uroEHx8fnC0tLVcHBwcAAAAAAFlZWQDj4+MDMzMzQ0ZGRiwtLS1nHR0dtkFBQR8rKysAAAAAAP///wAFBQUAOjo6WSUlJZyvr68Eg4ODAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlpaWAP///wEnJyePJiYmcQAAAAD///8AXV1dAH19fQAVFRUAQEBAMB4eHr88PDxEAAAAAP///wDS0tIAAAAAAAAAAAAxMTFjJSUlmf///wOUlJQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADY2NgAAAAAAC0tLWodHR2Y////A5aWlgChoaEAcnJyAJKSkgohISGbJCQkf////wKNjY0AlZWVAODg4ADV1dUAAAAAACgoKHYdHR2GAAAAALW1tQDw8PAA09PTAMjIyADo6OgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAALGxsQAqKioAPDw8ORsbG7hTU1MaVFRUAP///wAAAAAAOTk5TyAgILNFRUUaOzs7AGdnZwDBwcEA6OjoAJiYmAD///8CKCgolCIiImoAAAAA////AQAAAAA9PT0Ao6OjAHBwcACcnJwAAAAAAAAAAAAAAAAAAAAAANvb2wDS0tIAISEhAGxsbABkZGQPHBwcsDY2NkwLCwsAT09PAFJSUg8iIiKoMjIyWwAAAAD///8A////AAAAAAArKysAbW1tAGxsbA8iIiKuNDQ0RUFBQU4lJSWNLi4uVEJCQhkAAAAAYWFhADQ0NADZ2dkAAAAAAN/f3wAAAAAAYGBgAIGBgQAgICAAyMjIBIeHhxIjIyN3JiYmlPDw8AQAAAAAKCgoVyAgIKtsbGwRYWFhAFxcXAAXFxcAAAAAAIqKigA+Pj4APz8/LCEhIblKSkohPj4+KCIiIncfHx+rHR0dtiwsLGWAgIAMQEBAACoqKgAAAAAAv7+/AAAAAABHR0cA////AT8/PzQnJyePMTExikZGRjkiIiK4RkZGJ3V1dQ4fHx+qLy8vVQAAAAAeHh4ABQUFAEJCQgAAAAAAHh4eAAgICAAoKChcJSUlopGRkQg7OzsAAAAAAICAgA06OjpIIyMjsCoqKoNvb28IXV1dAP///wAAAAAAICAgAFVVVRMjIyN6HBwcuyEhIYA6Ojonvr6+BScnJ5IoKChyOjo6Tx4eHqtRUVEPioqKBTg4OD4wMDBWTU1NGgAAAACRkZEA1dXVAx0dHZctLS1tAAAAAM7OzgDu7u4AdXV1ABISEgBQUFAkHBwctioqKj4TExMATk5OAD8/PwBYWFgXJSUlnCQkJKBHR0cw////ADg4OAAiIiIAPT09Ph4eHr0iIiK7LCwsYwAAAAAuLi5ZDAwM6QcHB/seHh6pdHR0Ek9PTwBJSUkiHBwcuUJCQjA0NDQAh4eHAM7OzgCsrKwAxMTEAP///wIiIiKaKysrXAAAAAB0dHQA////AicnJ34lJSWZZmZmESYmJgCbm5sAcXFxAHl5eQCpqakHGBgYoBQUFNNXV1ce////AxoaGqMBAQH/AAAA/wwMDOxHR0c2BgYGADAwMGEeHh6iiIiICHd3dwAAAAAAp6enAP///wAsLCwAS0tLKR4eHrkuLi40HR0dADMzMwAsLCwSHh4es0ZGRjo4ODgAAAAAALy8vADf398A9PT0AAAAAAAaGhp/ExMT2k1NTSgAAAAAJSUleAcHB/0CAgL/FhYWyVxcXB2NjY0JJSUlpCgoKFoAAAAA////AP///wBUVFQAAAAAAFBQUCMjIyOoKCgogI2NjQRjY2MAPj4+ADw8PA4fHx+uNzc3VAAAAAB0dHQA////ALu7uwBpaWkAa2trDx0dHasgICDBIyMjjJWVlQVXV1cTKSkpdSAgIJQ9PT07AAAAADU1NT8jIyOyS0tLGVpaWgBRUVEAAAAAAFtbWwstLS1THx8ftiYmJoRzc3MPSEhIAEBAQACenp4AAAAAADIyMlUgICC7MTExU5SUlAoAAAAAPj4+ADg4OAA+Pj43HBwcsU9PTzUcHByyLi4uVgAAAAAAAAAA////ARYWFgAAAAAAMTExazQ0NGoAAAAA5OTkBFFRUR8xMTFbICAgph8fH7ItLS1WdXV1B0JCQgAfHx8A5+fnAIODgwB4eHgA1dXVAzc3N1EiIiKzJCQkoi4uLltBQUEmioqKCTIyMmwhISGNAAAAADIyMkYhISG3SEhILAICAgAAAAAA5ubmA1VVVQ1VVVU2RERETC4uLmglJSWZICAguCQkJKYzMzNhX19fGAAAAABaWloALS0tAPLy8gAAAAAAubm5AG1tbQB8fHwAAAAAAElJSRokJCRkHR0dpRwcHLgfHx+qFhYW0R4eHrcwMDBwNjY2ax4eHskYGBi9KCgodxgYGIYeHh6bHh4erxoaGrgdHR2zHx8fmSMjI2k9PT0/cHBwFAAAAABVVVUA////AKmpqQCKiooAAAAAAAAAAAAAAAAAiYmJAIODgwDg4OAARkZGAAAAAACNjY0JSkpKJTg4OFUYGBjRLCwsoS4uLpYoKCibJSUlmhcXF8MZGRnkMjIygSwsLGE+Pj5HQkJCK2RkZBTn5+cEAAAAACwsLGA1NTVmAAAAAMPDwwBubm4A0tLSAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAKampgCurq4A8vLyAH5+fgBHR0cASkpKHyQkJLJRUVEhAAAAAP///wP///8BRkZGISAgILAwMDBpAAAAAC0tLQA9PT0AaWlpAJSUlAAUFBQAKCgoTyYmJqaFhYULenp6AGBgYAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAENDQwDW1tYAaGhoADg4OAA3Nzc1KSkpsGhoaBR3d3cAioqKAJ+fnwAeHh4APj4+NCAgILkwMDBZ////AHt7ewBmZmYA19fXAERERABAQEAiJCQks0xMTCROTk4As7OzAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA7OzsAICAgAB8fHz4kJCSpioqKDYiIiADPz88AAAAAABUVFQB0dHQPPDw8Th0dHbwyMjJV////AGhoaAAwMDAAc3NzAG9vbw4oKCipNDQ0PjAwMABsbGwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEFBQQAjIyMAIiIiPSUlJaqFhYUOhYWFAAAAAAApKSkASUlJHSgoKJtPT082MjIyRB8fH7stLS1lpqamByUlJQBxcXEAqqqqCCcnJ6MoKChGIyMjAFJSUgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAeHh4AD8/PwA9PT0uJiYms1VVVRhoaGgAMjIyAGJiYhUjIyOcJiYmjHl5eQsAAAAAPz8/OiEhIbElJSWSODg4MFdXVws8PDwuIyMjuUBAQDE9PT0Af39/AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAC+vr4AXFxcAFxcXBAjIyOtMDAwUv///wNCQkIsIiIioSIiIpdNTU0SNDQ0AMnJyQAEBAQAQ0NDHiMjI4UaGhq7Hh4esBsbG7woKCiBgYGBCGtrawD///8AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAALCwsAC0tLQAAAAAADc3N1YhISG7IyMjoiEhIbwrKyuCZmZmEjg4OAAAAAAAmpqaAP///wA2NjYA////A01NTSVCQkJJRkZGOXx8fAlFRUUAPDw8AMjIyAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwMDAJWVlQB3d3cAsLCwBDMzMzYiIiJXNjY2MqysrARCQkIAAwMDAMrKygDx8fEA////AGFhYQCQkJAAREREACEhIQAvLy8Aa2trAD09PQDPz88AAQEBAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAxMTEAGlpaQCAgIAALCwsAAwMDAAxMTEAcHBwAEpKSgD7+/sA19fXAAAAAAAAAAAA19fXAJmZmQC8vLwArKysAK6urgDn5+cAvLy8ANnZ2QAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAxMTEAP///wAjIyMAAAAAADs7OwD///8A2dnZAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA//8AP/gAAD/4AAA/+AAAP/gAAD/4AAA/+AAAP/gAAD/4AAAD+AAAAeAAIACAAAAAgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEAAAADgAAAB+AAAA/wAAAP/AAAD/wAAA/8AAAP/AAAD/wAAA/8AAAP/gBgH/8B//8=
"""
# 解码base64编码的图像文件
icon_data = base64.b64decode(base64_icon)
icon_image = Image.open(io.BytesIO(icon_data))
# 将图像转换为Tkinter的PhotoImage对象
tk_image = ImageTk.PhotoImage(icon_image)
# 设置窗口图标
root.iconphoto(True, tk_image)
root.title("Nuclei GUI-v1.0    by：ryuuz4k1")
#调用api设置成由应用程序缩放
ctypes.windll.shcore.SetProcessDpiAwareness(1)
# 调用api获得当前的缩放因子
ScaleFactor=ctypes.windll.shcore.GetScaleFactorForDevice(0)
# 设置缩放因子
# root.tk.call("tk", "scaling", ScaleFactor / 100)
# 获取屏幕宽高
screen_width = root.winfo_screenwidth() * ScaleFactor / 100
screen_height = root.winfo_screenheight()  * ScaleFactor / 100
root_width = 1600
root_height = 1200
# 设置窗口的位置
root.geometry(f"{int(root_width)}x{int(root_height)}+{int((screen_width - root_width) / 2)}+{int((screen_height - root_height) / 4)}")
# 设置窗口不可调整大小
root.resizable(False, False)
if not os.path.exists(POC_FOLDER):
    os.makedirs(POC_FOLDER)

if not os.path.exists(WORK_FOLDER):
    os.makedirs(WORK_FOLDER)

config = configparser.ConfigParser()
config.read(CONFIG_FILE_PATH)

# 检查配置文件是否存在，如果不存在则创建
if not config.read(CONFIG_FILE_PATH):
    # 如果配置文件不存在，则创建默认的配置
    config["Settings"] = {
        "User-Agent": "",
        "Proxy": "",
        "Random-UA": "False",
        "API_Key": "",
        "cloud_upload": "False"
    }
    # 将默认配置写入配置文件
    with open(CONFIG_FILE_PATH, "w", encoding="utf-8") as configfile:
        config.write(configfile)

# 保存配置到配置文件
def save_config():
    global CONFIG_FILE_PATH

    with open(CONFIG_FILE_PATH, "w", encoding="utf-8") as configfile:
        config.write(configfile)

UA = config.get("Settings", "User-Agent")
RANDOM_UA_STATUS = config.getboolean("Settings", "Random-UA")
PROXY = config.get("Settings", "Proxy")
API_KEY= config.get("Settings", "API_Key")
CLOUD_UPLOAD_STATUS = config.getboolean("Settings", "cloud_upload")

# 设置UA头
def set_ua():
    global UA
    global RANDOM_UA_STATUS
    
    def save_ua_config():
        global UA

        UA = ua_entry.get()
        config.set("Settings", "User-Agent", UA)
        save_config()
    
    def save_random_ua_config():
        global UA
        global RANDOM_UA
        global RANDOM_UA_STATUS
        global WORK_FOLDER

        RANDOM_UA_STATUS = random_ua_button_var.get()
        config.set("Settings", "Random-UA", str(RANDOM_UA_STATUS))  # 将布尔值转换为字符串保存
        save_config()
        random_ua_path = f"{WORK_FOLDER}/random_UA.txt"
        if not os.path.exists(random_ua_path):
            random_UA = """Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/118.0
Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36
Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; Touch; rv:11.0) like Gecko
Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36 Edg/114.0.1788.0
Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.25 Safari/537.36 Core/1.70.3883.400 QQBrowser/10.8.4559.400
Mozilla/5.0 (Macintosh; Intel Mac OS X 11_6) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.5 Safari/605.1.15
Mozilla/5.0 (X11; Linux x86_64; rv:109.0) Gecko/20100101 Firefox/114.0
Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36 uacq
Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:109.0) Gecko/20100101 Firefox/112.0
Mozilla/5.0 (X11; U; Linux sparc64; es-PY; rv:5.0) Gecko/20100101 IceCat/5.0 (like Firefox/5.0; Debian-6.0.1)"""

            with open(random_ua_path, "w", encoding="utf-8") as file:
                file.write(random_UA)

        else:
            if RANDOM_UA_STATUS:
                with open(random_ua_path, "r", encoding="utf-8") as file:
                    lines = file.readlines()
                    choose_UA = random.choice(lines)
                    RANDOM_UA = choose_UA.strip()
                    print(RANDOM_UA)

    # 创建一个Toplevel窗口，即弹窗
    ua_window = tk.Toplevel(root)
    ua_window.title("设置UA头")
    ua_window.resizable(False, False)
    x = root.winfo_rootx() + 400
    y = root.winfo_rooty() + 80
    ua_window.geometry(f"600x280+{x}+{y}")

    # 在弹窗中添加Label和Entry，用于输入UA头
    ua_label = ttk.Label(ua_window, font=("黑体", 18), text="请输入User-Agent:")
    ua_label.place(relx=0.04, rely=0.1, relwidth=0.89)
    ua_entry = tk.Entry(ua_window, font=("黑体", 20))
    ua_entry.insert(0, UA)
    ua_entry.place(relx=0.044, rely=0.3, relwidth=0.894)

    def random_ua():
        global RANDOM_UA_STATUS

        RANDOM_UA_STATUS = random_ua_button_var.get()
        
        if RANDOM_UA_STATUS:
            ua_delete_canvas.place(relx=0.02, rely=0.36, relwidth=0.95)
        else:
            ua_delete_canvas.place(relx=1, rely=1, relwidth=0)

    random_ua_button_style = ttk.Style()
    random_ua_button_style.configure("Custom.TCheckbutton", font=("黑体", 14, ""))
    random_ua_button_var = tk.BooleanVar()
    random_ua_button_var.set(RANDOM_UA_STATUS)
    random_ua_button = ttk.Checkbutton(ua_window, text="使用随机UA头", variable=random_ua_button_var, command=random_ua, style="Custom.TCheckbutton")
    random_ua_button.place(relx=0.04, rely=0.42, relwidth=0.4, relheight=0.2)
    ua_delete_canvas = tk.Canvas(ua_window, height=2, highlightthickness=0, bg="black")

    if RANDOM_UA_STATUS:
        ua_delete_canvas.place(relx=0.02, rely=0.36, relwidth=0.95)

    # 在弹窗中添加"确定"按钮，点击后获取Entry中的值
    def ok_button_clicked():
        save_ua_config()
        save_random_ua_config()
        ua_window.destroy()
    
    def cancel_button_clicked():
        ua_window.destroy()

    ok_button = ttk.Button(ua_window, text="确定", command=ok_button_clicked, style="Custom.TButton")
    ok_button.place(relx=0.04, rely=0.64, relwidth=0.24, relheight=0.22)
    cancel_button = ttk.Button(ua_window, text="取消", command=cancel_button_clicked, style="Custom.TButton")
    cancel_button.place(relx=0.7, rely=0.64, relwidth=0.24, relheight=0.22)
    # 设置焦点并阻塞主窗口
    ua_window.focus_force()
    ua_window.grab_set()
    root.wait_window(ua_window)
    # 释放焦点
    ua_window.grab_release()

# 设置代理
def set_proxy():
    global PROXY
    
    def save_proxy_config():
        global PROXY

        PROXY = proxy_entry.get()
        config.set("Settings", "Proxy", PROXY)
        save_config()

    # 创建一个Toplevel窗口，即弹窗
    proxy_window = tk.Toplevel(root)
    proxy_window.title("设置代理")
    proxy_window.resizable(False, False)
    x = root.winfo_rootx() + 400
    y = root.winfo_rooty() + 80
    proxy_window.geometry(f"600x280+{x}+{y}")

    # 在弹窗中添加Label和Entry，用于输入代理地址和端口
    proxy_label = ttk.Label(proxy_window, font=("黑体", 18), text="请输入代理地址和端口:")
    proxy_label.place(relx=0.04, rely=0.1, relwidth=0.89) 
    proxy_entry = tk.Entry(proxy_window, font=("黑体", 20))
    proxy_entry.insert(0, PROXY)
    proxy_entry.place(relx=0.044, rely=0.36, relwidth=0.894)
    
    # 在弹窗中添加"确定"按钮，点击后获取Entry中的值
    def ok_button_clicked():
        save_proxy_config()
        update_proxy_status()
        proxy_window.destroy()
    
    def cancel_button_clicked():
        proxy_window.destroy()

    ok_button = ttk.Button(proxy_window, text="确定", command=ok_button_clicked, style="Custom.TButton")
    ok_button.place(relx=0.04, rely=0.64, relwidth=0.24, relheight=0.22)
    cancel_button = ttk.Button(proxy_window, text="取消", command=cancel_button_clicked, style="Custom.TButton")
    cancel_button.place(relx=0.7, rely=0.64, relwidth=0.24, relheight=0.22)
    
    # 设置焦点并阻塞主窗口
    proxy_window.focus_force()
    proxy_window.grab_set()
    root.wait_window(proxy_window)
    # 释放焦点
    proxy_window.grab_release()

# 更新代理状态标签
def update_proxy_status():
    global PROXY

    if PROXY:
        proxy_status_label.config(text=f"代理生效中 → {PROXY}", foreground="blue", font=("黑体", 20, "bold"))

    else:
        proxy_status_label.config(text="当前未设置代理", foreground="black", font=("黑体", 20, "bold"))

# 执行Nuclei扫描
def run_scan(target_url):
    global UA
    global API_KEY
    global RANDOM_UA
    global POC_FOLDER
    global WORK_FOLDER
    global ERROR_MESSAGE
    global MANUALLY_STOP
    global SELECT_ALL_POC
    global SELECTED_FILES
    global RANDOM_UA_STATUS
    global CLOUD_UPLOAD_STATUS
    global CURRENT_SCAN_STATUS
    global CURRENT_SCAN_PROCESS
    global CURRENT_POCLIST_RESULT
    global CURRENT_VALIDATE_RESULT
    global CURRENT_VALIDATE_STATUS

    MANUALLY_STOP = False
    command = [NUCLEI_PATH, "-o", "./work/result.txt"]
    file_path_new = os.path.join(WORK_FOLDER, "result_new.txt")
    file_path_zh = os.path.join(WORK_FOLDER, "result_zh.txt")
    file_path = os.path.join(WORK_FOLDER, "poc_list.xlsx")

    # 检查是否有扫描正在进行中
    if CURRENT_VALIDATE_STATUS or CURRENT_SCAN_STATUS:
        show_scan_status("请稍等，正在扫描中...", "blue")
        return
    
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    # 创建ID到名称的映射字典
    id_to_name = {}

    for row in sheet.iter_rows(values_only=True):
            id_to_name[row[0]] = row[1]

    if not os.path.exists(NUCLEI_PATH):
        show_scan_status("请先加载Nuclei！", "red")
        print("不存在nuclei.exe文件")
        wb.close()
        return
    
    open_xlsx = update_poclist()
    
    if open_xlsx == 1:
        return

    # 更新表格
    def update_table(data):
        if data:
            # 插入新数据并替换名称
            for item in data:
                name, *rest = item

                # 如果名称存在于映射字典中，则替换为对应的名称
                if name in id_to_name:
                    name = id_to_name[name]

                cleaned_item = [name, *rest]
                zh_result = [cleaned_item[0].strip("[]"), cleaned_item[1].strip("[]"), cleaned_item[2]]
                file_zh.write("\t".join(zh_result) + "\n")
                new_row_id = table.insert("", "end", values=cleaned_item, tags=("center"))
                tag = "even" if table.index(new_row_id) % 2 == 0 else ""
                table.item(new_row_id, tag=tag)

        else:
            for row_id in table.get_children():
                table.delete(row_id)

    # 首先清空表格
    update_table([])
    scan_result = []

    with open(file_path_new, "w", encoding="utf-8") as file:
        file.write("")
    
    with open(file_path_zh, "w", encoding="utf-8") as file:
        file.write("")

    if not SELECTED_FILES:
        show_scan_status("请先选择POC！", "red")
        wb.close()
        return
    
    else:
        if not target_url and not URL_FILE_PATH:
            show_scan_status("请先输入URL！", "red")
            wb.close()
            return
    
    if not CURRENT_POCLIST_RESULT:
        show_scan_status("部分POC错误，请先完善！", "red")
        wb.close()
        messagebox.showerror("错误", ERROR_MESSAGE)
        return
    
    show_scan_status("检查已选POC格式中...", "blue")
    CURRENT_VALIDATE_STATUS = True
    validate_poc()
    
    if CURRENT_VALIDATE_STATUS or MANUALLY_STOP:
        wb.close()
        return
    
    if not CURRENT_VALIDATE_RESULT:
        wb.close()
        show_scan_status("已选POC格式错误，扫描已终止", "red")
        messagebox.showerror("错误", ERROR_MESSAGE)
        return

    show_scan_status("请稍等，正在扫描中...", "blue")
    CURRENT_SCAN_STATUS = True

    if CLOUD_UPLOAD_STATUS:
        command.extend(["-cup"])

    if PROXY:
        command.extend(["-p", PROXY])

    if RANDOM_UA_STATUS:
        random_ua_path = f"{WORK_FOLDER}/random_UA.txt"

        with open(random_ua_path, "r", encoding="utf-8") as file:
                lines = file.readlines()
                choose_UA = random.choice(lines)
                RANDOM_UA = choose_UA.strip()

        command.extend(["-H", "User-Agent: " + RANDOM_UA])
    
    else:
        if UA:
            command.extend(["-H", "User-Agent: " + UA])

    if SELECT_ALL_POC:
        command.extend(["-t", POC_FOLDER])

    else:
        if SELECTED_FILES:  # 如果选择了POC文件，则执行-t参数
            selected_files = ["./poc/{}".format(file) for file in SELECTED_FILES]
            command.extend(["-t", ",".join(selected_files)])

    if target_url:
            command.extend(["-u", target_url])
    
    else:
        if URL_FILE_PATH:
            command.extend(["-l", URL_FILE_PATH])

    try:
        print("Executed command:", " ".join(command))
        CURRENT_SCAN_PROCESS = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, creationflags=subprocess.CREATE_NO_WINDOW)
        # 编译正则表达式以匹配 ANSI 转义码
        ansi_escape = re.compile(r"\x1b\[(?:\d+;)*\d+m")
        results = ""

        with open(file_path_new, "a", encoding="utf-8") as file:
            with open(file_path_zh, "a", encoding="utf-8") as file_zh:
                # 循环读取子进程的输出
                while True:
                    # 逐行读取输出
                    output = CURRENT_SCAN_PROCESS.stdout.readline()
                    print(output.strip("\n"))
                    results = results + output

                    if "[\x1b[92m" in output:
                        # 使用正则表达式替换 ANSI 转义码
                        output = ansi_escape.sub("", output)
                        parts = output.split()

                        if len(parts) == 4:
                            scan_result = [parts[0].strip("[]"), parts[2].strip("[]"), parts[3]]
                            file.write("\t".join(scan_result) + "\n")
                            update_table([[scan_result[0], scan_result[1], scan_result[2]]])

                    # 如果输出为空或子进程已经结束，则退出循环
                    if output == "" or CURRENT_SCAN_PROCESS.poll() is not None:
                        CURRENT_SCAN_STATUS = False
                        break

        # 检查代理是否可达
        if "all proxies are dead" in results:
            show_scan_status("代理地址访问失败！扫描已终止", "red")
            CURRENT_SCAN_STATUS = False
            wb.close()
            return
        
        if CURRENT_SCAN_STATUS or MANUALLY_STOP:
            wb.close()
            return
        
        if scan_result:
            show_scan_status("扫描结束，请验收扫描结果", "green")
        
        else:
            show_scan_status("扫描结束，未发现成功目标(T⌒T)", "green")

    except subprocess.CalledProcessError as e:
        print(f"Error executing nuclei: {e.output}")
        wb.close()
        
    except Exception as e:
        print(f"Unexpected error: {e}")
        show_scan_status("发生错误，扫描已终止", "red")
        error_lines = str(e).split("\n")
        second_last_line = error_lines[-1]
        first_colon_index = second_last_line.find("] ")
        content_after_last_colon = second_last_line[first_colon_index + 1:].strip()
        CURRENT_SCAN_STATUS = False
        messagebox.showerror("错误", content_after_last_colon)
        wb.close()
    
    CURRENT_SCAN_STATUS = False
    MANUALLY_STOP = False
    wb.close()

    if "could not upload results got status code 401" in results:
            messagebox.showerror("错误", "扫描结果上传云平台失败，云平台API_Key填写错误！")

# 异步执行扫描
def Scan(target_url):
    threading.Thread(target=run_scan, args=(target_url,)).start()

# 终止扫描按钮点击事件
def stop_scan():
    global MANUALLY_STOP
    global CURRENT_SCAN_STATUS
    global CURRENT_SCAN_PROCESS
    global CURRENT_VALIDATE_STATUS
    global CURRENT_VALIDATE_PROCESS

    MANUALLY_STOP = True

    if CURRENT_VALIDATE_STATUS:
        if CURRENT_VALIDATE_PROCESS:
            CURRENT_VALIDATE_PROCESS.terminate()
            show_scan_status("扫描已手动终止", "blue")
            CURRENT_VALIDATE_STATUS = False

    elif CURRENT_SCAN_STATUS:
        if CURRENT_SCAN_PROCESS:
            CURRENT_SCAN_PROCESS.terminate()
            show_scan_status("扫描已手动终止", "blue")
            CURRENT_SCAN_STATUS = False

    else:
        show_scan_status("当前没有运行中的扫描！", "red")

# 检查Nuclei加载状态
def load_nuclei():
    global NUCLEI_PATH
    global CURRENT_LOAD_STATUS 

    try:
        if os.path.exists(NUCLEI_PATH):
            command = [NUCLEI_PATH, "--version"]
            print("Executed command:", " ".join(command))
            process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, creationflags=subprocess.CREATE_NO_WINDOW)
            result, _ = process.communicate()
            print(result)

            # 检查输出中是否包含 "Nuclei Engine Version" 字符串
            if "Nuclei Engine Version" in result:
                show_nuclei_status("Nuclei加载成功( •̀ ω •́ )", "green")
                # 解析版本号
                lines = result.split("\n")  # 拆分字符串成行
                first_line = lines[0]  # 获取第一行
                colon_index = first_line.find(": ")
                    
                if colon_index != -1:
                    version = first_line[colon_index + 1:].strip()
                    show_nuclei_version("Nuclei版本：" + version, "black")

        else:
            show_nuclei_version("", "black")
            show_nuclei_status("Nuclei加载失败！", "red")
            print("nuclei.exe不存在")
            CURRENT_LOAD_STATUS = False
            return
            
    except Exception as e:
        print(f"Unexpected error: {e}")
        show_nuclei_version("", "black")
        show_nuclei_status("Nuclei加载失败！", "red")
        CURRENT_LOAD_STATUS = False
        return
    
    CURRENT_LOAD_STATUS = False

# 异步执行加载Nuclei
def Load():
    global CURRENT_LOAD_STATUS

    if CURRENT_LOAD_STATUS:
        return
    
    show_nuclei_version("", "black")
    show_nuclei_status("Nuclei加载中...", "blue")
    CURRENT_LOAD_STATUS = True
    threading.Timer(0.2, load_nuclei).start()
    
# 显示Nuclei状态
def show_nuclei_status(message, color):
    nuclei_status_label.config(text=message, foreground=color, font=("黑体", 20, "bold"))

# 显示Nuclei版本
def show_nuclei_version(message, color):
    nuclei_version_label.config(text=message, foreground=color, font=("黑体", 20, "bold"))

# 显示Nuclei logo
def show_nuclei_logo(message, color):
    nuclei_logo_label.config(text=message, foreground=color, font=("黑体", 20,))

# 编辑POC
def view_poc():
    global SELECTED_FILES
    global POC_FOLDER

    try:
        if SELECTED_FILES:
            show_validation_status("", "red")

            for file in SELECTED_FILES:
                os.system(f"start {POC_FOLDER}\{file}")

        else:
            show_validation_status("请先选择POC！", "red")

    except Exception as e:
        print(f"Unexpected error: {e}")
        return

# 选择POC
def select_poc():
    global POC_FOLDER
    global YAML_FILES
    global ERROR_MESSAGE
    global SELECTED_FILES
    global CURRENT_POCLIST_RESULT

    show_selectpoc_status("当前未选择POC！", "red")
    open_xlsx = update_poclist()

    if open_xlsx == 1:
        return

    if not YAML_FILES:
        show_selectpoc_status("未发现任何POC！", "red")
        return
    
    if not CURRENT_POCLIST_RESULT:
        show_selectpoc_status("部分POC错误，请先完善！", "red")
        messagebox.showerror("错误", ERROR_MESSAGE)
        return
    
    files = os.listdir("poc")  # 获取poc文件夹内的文件列表
    poc_window = tk.Toplevel(root)  # 创建弹窗
    poc_window.title("选择POC")
    poc_window.resizable(False, False)
    x = root.winfo_rootx() + 400
    y = root.winfo_rooty() + 160
    poc_window.geometry(f"826x900+{x}+{y}")

    columns = ("Severity", "Name")
    file_tree = ttk.Treeview(poc_window, columns=columns, show="headings", height=20, style="Custom.Treeview")

    def file_tree_color(): # 表格栏隔行显示不同颜色函数
        items = file_tree.get_children()
        for index, iid in enumerate(items):
            tag = "even" if index % 2 == 0 else ""
            file_tree.item(iid, tag=tag)

    def sort_column(tree, col, reverse):
        data = [(tree.set(child, col), child) for child in tree.get_children("")]

        if col == columns[0]:  # 检查是否是第一列
            custom_order = {"info": 0, "low": 1, "medium": 2, "high": 3, "critical": 4, "unknown": 5}
            data.sort(key=lambda x: custom_order.get(x[0], float("inf")), reverse=reverse)
        else:
            # 使用 locale 模块设置地区设置为中国简体中文
            locale.setlocale(locale.LC_COLLATE, "zh_CN.UTF-8")
            data.sort(key=lambda x: locale.strxfrm(x[0]), reverse=reverse)

        for index, item in enumerate(data):
            tree.move(item[1], "", index)

        tree.heading(col, command=lambda: sort_column(tree, col, not reverse))
        file_tree_color()

    # 设置表头
    for col in columns:
        file_tree.heading(col, text=col, anchor="center", command=lambda c=col: sort_column(file_tree, c, False))
        file_tree.bind("<B1-Motion>", lambda e: "break")

    # 设置表格布局
    file_tree.column("Severity", width=140)
    file_tree.column("Name", width=620)
    file_tree.place(relx=0.03, rely=0.014)
    file_tree.tag_configure("even", background=light_gray)

    # 在表格内添加边框
    file_tree_canvas_x = tk.Canvas(poc_window, height=1, highlightthickness=0, bg="gray")
    file_tree_canvas_x.place(relx=0.032, rely=0.043, relwidth=0.92)
    file_tree_canvas_y = tk.Canvas(poc_window, width=1, highlightthickness=0, bg="gray")
    file_tree_canvas_y.place(relx=0.2, rely=0.014, relheight=0.875)
   
    # 将字段值添加到表格
    for file in files:
        if file.lower().endswith((".yaml")):
            yaml_path = os.path.join(POC_FOLDER, file)

        with open(yaml_path, "r", encoding="utf-8") as f:
            yaml_data = f.read()
            severity_value = ""
            name_value = ""

            # 按行解析字段值
            for line in yaml_data.split("\n"):
                if "severity:" in line:
                    severity_value = line.split(":")[1].strip()
                elif "name:" in line:
                    name_value = line.split(":")[1].strip()

            if name_value:
                file_tree.insert("", tk.END, text=file, values=(severity_value, name_value))
    
    # 添加完字段值后立即按漏洞名称列排序
    sort_column(file_tree, "Name", False)

    # 设置滚动条与表格的关联
    y_scrollbar = ttk.Scrollbar(poc_window, orient="vertical", command=file_tree.yview)
    y_scrollbar.place(relx=0.95, rely=0.023, relheight=0.886)
    y_scrollbar.lower()
    file_tree.configure(yscrollcommand=y_scrollbar.set)
    cover_label = ttk.Label(poc_window, text="")
    cover_label.place(relx=0.95, rely=0.889, relheight=0.02)
    
    file_path = os.path.join(WORK_FOLDER, "poc_list.xlsx")
    wb = openpyxl.load_workbook(file_path)

    # 更新选中的文件名
    def update_selected_files(event):
        global WORK_FOLDER
        global SELECTED_FILES
        global TREEVIEW_SELECT
        
        if not os.path.exists(file_path):
            print("不存在poc_list.xlsx文件")
            poc_window.destroy()
            messagebox.showerror("错误", "请先更新POC列表！")
            return
        sheet = wb.active
        TREEVIEW_SELECT = True
        SELECTED_FILES.clear()
        selected_items = file_tree.selection()

        for item in selected_items:
            SELECTED_FILES.append(file_tree.item(item, "values")[1])
        
        # 遍历SELECTED_FILES列表
        for index, name in enumerate(SELECTED_FILES):
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=4):
                current_name = row[0].value  # 获取当前行的name
                current_filename = row[2].value  # 获取当前行的文件名
                # 如果当前行的name和SELECTED_FILES中的name匹配
                if current_name == name:
                    # 替换SELECTED_FILES中的元素为对应的文件名
                    SELECTED_FILES[index] = current_filename
                    break  # 找到匹配的就可以退出内层循环

        # 更新全选复选框的状态
        items = file_tree.get_children()
        if len(file_tree.selection()) == len(items):
            select_all_button_var.set(True)
        else:
            select_all_button_var.set(False)

    # 绑定事件，当用户选择文件时更新选中的文件名
    file_tree.bind("<<TreeviewSelect>>", update_selected_files)
    
    def select_and_close(event):
        # 检查是否按住了 Ctrl 或 Shift 键，如果是则直接返回，不执行操作
        if event.state & 4 or event.state & 1:
            return
            
        if event.y > 24:
            update_selected_files(None)  # 更新选中的文件名
            close_window()  # 关闭窗口

    file_tree.bind("<Double-1>", select_and_close)
    
    # 点击确定按钮后关闭弹窗
    def close_window():
        global SELECT_ALL_POC
        global SELECTED_FILES
        global TREEVIEW_SELECT

        if len(SELECTED_FILES) >= 1:
            show_selectpoc_status("已选择" + str(len(SELECTED_FILES)) + "个POC", "green")

        else:
            show_selectpoc_status("当前未选择POC！", "red")

        # 判断是否触发了 <<TreeviewSelect>> 事件
        if not TREEVIEW_SELECT:
            SELECTED_FILES.clear()
            show_selectpoc_status("当前未选择POC！", "red")

        poc_window.destroy()
        TREEVIEW_SELECT = False
        SELECT_ALL_POC = False
        if select_all_button_var.get():
            SELECT_ALL_POC = True

    # 全选功能
    def select_all():
        items = file_tree.get_children()
        if len(file_tree.selection()) == len(items):  # 判断是否已全选
            select_all_button_var.set(False)  # 如果已全选，勾选复选框
        else:
            select_all_button_var.set(True)  # 如果未全选，取消复选框的勾选状态

        if len(file_tree.selection()) == len(items):  # 如果未全选，执行全选操作
            for item in items:
                file_tree.selection_remove(item)
        else:
            for item in items:
                file_tree.selection_add(item)

    select_all_button_style = ttk.Style()
    select_all_button_style.configure("Custom.TCheckbutton", font=("黑体", 14, ""))
    select_all_button_var = tk.BooleanVar()
    select_all_button = ttk.Checkbutton(poc_window, text="全选", variable=select_all_button_var, command=select_all, style="Custom.TCheckbutton")
    select_all_button.place(relx=0.03, rely=0.89, relwidth=0.15, relheight=0.06)
    confirm_button = ttk.Button(poc_window, text="确定", command=close_window, style="Custom.TButton")
    confirm_button.place(relx=0.4, rely=0.91, relwidth=0.2, relheight=0.06)
    
    def close():
        global SELECTED_FILES

        SELECTED_FILES.clear()
        show_selectpoc_status("当前未选择POC！", "red")
        poc_window.destroy()

    poc_window.protocol("WM_DELETE_WINDOW", close)
    # 设置焦点并阻塞主窗口
    poc_window.focus_force()
    poc_window.grab_set()
    root.wait_window(poc_window)
    # 释放焦点
    poc_window.grab_release()
    wb.close()
    print(SELECTED_FILES)
    return SELECTED_FILES
    
#选择POC状态
def show_selectpoc_status(message, color):
    selectpoc_status_label.config(text=message, foreground=color, font=("黑体", 20, "bold"))

# 编辑URL
def view_urls():
    global URL_FILE_PATH

    try:
        if URL_FILE_PATH:
            os.system(f"start {URL_FILE_PATH}")

        else:
            show_urls_status("请先选择URL文件！", "red")

    except Exception as e:
        print(f"Unexpected error: {e}")
        return

# 显示编辑状态
def show_urls_status(message, color):
    urls_status_label.config(text=message, foreground=color, font=("黑体", 20, "bold"))

# 选择URL
def select_url_file():
    global URL_FILE_PATH

    last_selected_path = URL_FILE_PATH
    URL_FILE_PATH = filedialog.askopenfilename(initialdir=os.getcwd(), filetypes=[("Text Files", "*.txt")])
    
    if URL_FILE_PATH:
        urls_status_label.place(relx=1, rely=1, relwidth=0, relheight=0)
        url_file_entry.config(state=tk.NORMAL, foreground="black")
        url_file_entry.delete(0, tk.END)
        url_file_entry.insert(0, URL_FILE_PATH)
        url_file_entry.config(state="disabled")
        url_content = target_url_entry.get().strip()
        print(f"[{URL_FILE_PATH}]")

        if url_content:
            url_delete_canvas.place(relx=0.249, rely=0.206, relwidth=0.46)

    else:
        URL_FILE_PATH = last_selected_path

# 新建POC
def new_poc():
    global POC_FOLDER

    try:
        # 获取当前时间戳
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        # 构建新的POC文件路径
        new_poc_path = os.path.join(POC_FOLDER, f"poc_{timestamp}.yaml")

        # 创建新的POC文件
        with open(new_poc_path, "w", encoding="utf-8") as new_poc_file:
            # 添加默认内容
            new_poc_file.write(f"id: poc_{timestamp}\n\n")
            new_poc_file.write("info:\n")
            new_poc_file.write("  name: \n")
            new_poc_file.write("  author: \n")
            new_poc_file.write("  severity: \n")
            new_poc_file.write("  tags: \n\n")
            new_poc_file.write("http:\n")
            new_poc_file.write("  - raw:\n")
            new_poc_file.write("      - |\n")
            new_poc_file.write("        \n\n")
            new_poc_file.write("    matchers:\n")
            new_poc_file.write("      - type: dsl\n")
            new_poc_file.write("        dsl:\n")
            new_poc_file.write("          - ")

        # 使用系统默认编辑器打开新建的POC文件
        os.system(f"start {new_poc_path}")

    except Exception as e:
        print(f"Unexpected error: {e}")
        return

# 检查POC
def validate_poc():
    global NUCLEI_PATH
    global ERROR_MESSAGE
    global SELECTED_FILES
    global CURRENT_VALIDATE_STATUS
    global CURRENT_VALIDATE_RESULT
    global CURRENT_VALIDATE_PROCESS

    CURRENT_VALIDATE_STATUS = True
    CURRENT_VALIDATE_RESULT = False
    
    selected_files = ['"./poc/{}"'.format(file) for file in SELECTED_FILES]
    command = [NUCLEI_PATH, "-t", ",".join(selected_files), "-validate"]
    
    try:
        print("Executed command:", " ".join(command))
        CURRENT_VALIDATE_PROCESS = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, creationflags=subprocess.CREATE_NO_WINDOW)
        result, _ = CURRENT_VALIDATE_PROCESS.communicate()
        print(result)

        if "All templates validated successfully" in result:
            CURRENT_VALIDATE_RESULT = True
        
        else:
            result_lines = result.split("\n")
            error_result = ""

            for line in result_lines:
                if "ERR" in line:
                    yaml_index = line.rfind(".yaml:")
                    backslash_index = line[:yaml_index].rfind("\\")
                    error_result += line[backslash_index + 1:] + "\n\n"

            ERROR_MESSAGE = error_result.strip()

    except Exception as e:
        print(f"Unexpected error: {e}")
        show_nuclei_version("", "black")
        show_nuclei_status("Nuclei加载失败！", "red")

    CURRENT_VALIDATE_STATUS = False

# 显示验证结果
def show_validation_status(message, color):
    validation_status_label.config(text=message, foreground=color, font=("黑体", 20, "bold"))

# 显示扫描状态
def show_scan_status(message, color):
    scan_status_label.config(text=message, foreground=color, font=("黑体", 20, "bold"))

# 查看POC列表
def view_pocs():
    global WORK_FOLDER

    file_path = os.path.join(WORK_FOLDER, "poc_list.xlsx")

    if not os.path.exists(file_path):
        print("不存在poc_list.xlsx文件")
        messagebox.showerror("错误", "请先更新POC列表！")
        return

    os.startfile(file_path)

# 查看扫描结果
def view_results():
    global WORK_FOLDER
    global CURRENT_SCAN_STATUS

    file_path_zh = os.path.join(WORK_FOLDER, "result_zh.txt")


    if CURRENT_SCAN_STATUS:
        messagebox.showerror("错误", "请等待当前扫描结束！")

    else:
        try:
            os.startfile(file_path_zh)

        except Exception as e:
            print(f"Unexpected error: {e}")
            messagebox.showerror("错误", "当前没有扫描结果！")
            return

# 定义表头排序逻辑
def sort_treeview(tree, col):
    current_state = sort_states[col]
    data = [(tree.set(child, col), child) for child in tree.get_children("")]

    if col == columns[1]:  # 检查是否是第二列
        custom_order = {"info": 0, "low": 1, "medium": 2, "high": 3, "critical": 4, "unknown": 5}
        data.sort(key=lambda x: custom_order.get(x[0], float("inf")), reverse=current_state)

    else:
        # 临时设置中文排序规则
        data.sort(key=lambda x: locale.strxfrm(x[0]), reverse=current_state)

    for i, item in enumerate(data):
        tree.move(item[1], "", i)

    sort_states[col] = not current_state
    table_color()

# 更新POC列表
def update_poclist():
    global POC_FOLDER
    global YAML_FILES
    global WORK_FOLDER
    global ERROR_MESSAGE
    global CURRENT_POCLIST_RESULT

    locale.setlocale(locale.LC_ALL, "zh_CN.UTF-8")
    CURRENT_POCLIST_RESULT = False

    # 检查POC文件夹是否存在
    if os.path.exists(POC_FOLDER):
        # 获取POC文件夹内所有yaml文件的文件名列表
        yaml_files1 = [file for file in os.listdir(POC_FOLDER) if file.endswith(".yaml")]
        YAML_FILES = sorted(yaml_files1, key=locale.strxfrm)

        # 创建一个新的Excel工作簿
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "poc列表"
        # 设置表头
        bold_font = Font(bold=True)
        sheet["A1"] = "id"
        sheet["A1"].font = bold_font
        sheet["B1"] = "name"
        sheet["B1"].font = bold_font
        sheet["C1"] = "severity"
        sheet["C1"].font = bold_font
        sheet["D1"] = "文件名"
        sheet["D1"].font = bold_font
        row_index = 2  # 从第二行开始写入数据
        sheet.auto_filter.ref = sheet.dimensions
        sheet.column_dimensions["C"].width = 11.38
        # 存储已经写入的字段值
        written_ids = set()
        written_names = set()
        # 存储错误的文件名
        error_id = []
        error_name = []
        error_severity = []
        id_value = ""
        name_value = ""
        severity_value = ""
        filename_value = ""

        # 遍历yaml文件列表
        for yaml_file in YAML_FILES:
            if yaml_file.lower().endswith((".yaml")):
                yaml_path = os.path.join(POC_FOLDER, yaml_file)

            # 打开yaml文件并读取内容，获取字段值
            with open(yaml_path, "r", encoding="utf-8") as file:
                try:
                    yaml_content = file.read()

                    for line in yaml_content.split("\n"):
                        if "id:" in line:
                            id_value = line.split(":")[1].strip()
                                
                        elif "name:" in line:
                            name_value = line.split(":")[1].strip()
                                
                        elif "severity:" in line:
                            severity_value = line.split(":")[1].strip()
                                
                        filename_value = os.path.basename(file.name)
        
                    # 将字段值写入Excel表格
                    sheet[f"A{row_index}"] = id_value
                    sheet[f"B{row_index}"] = name_value
                    sheet[f"C{row_index}"] = severity_value
                    sheet[f"D{row_index}"] = filename_value
                    row_index += 1

                    # 设置表格内容左对齐
                    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                        for cell in row:
                            cell.alignment = Alignment(horizontal="left")

                    # 添加字段值到集合中
                    written_ids.add(id_value)
                    written_names.add(name_value)

                    if not id_value:
                        if not name_value:
                            if not severity_value:
                                error_id.append(yaml_file)
                                error_name.append(yaml_file)
                                error_severity.append(yaml_file)

                            else:
                                error_id.append(yaml_file)
                                error_name.append(yaml_file)

                        else:
                            if not severity_value:
                                error_id.append(yaml_file)
                                error_severity.append(yaml_file)

                            else:
                                error_id.append(yaml_file)

                    else:
                        if not name_value:
                            if not severity_value:
                                error_name.append(yaml_file)
                                error_severity.append(yaml_file)

                            else:
                                error_name.append(yaml_file)
                        
                        else:
                            if not severity_value:
                                error_severity.append(yaml_file)

                except OSError as e:
                    print(f"Unexpected error: {e}")
                    error_lines = str(e).split("\n")
                    second_last_line = error_lines[-1]
                    first_colon_index = second_last_line.find("] ")
                    content_after_last_colon = second_last_line[first_colon_index + 1:].strip()
                    wb.close()
                    messagebox.showerror("错误", content_after_last_colon)
                    return
                
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    error_lines = str(e).split("\n")
                    second_last_line = error_lines[-1]
                    first_colon_index = second_last_line.find("] ")
                    content_after_last_colon = second_last_line[first_colon_index + 1:].strip()
                    wb.close()
                    messagebox.showerror("错误", content_after_last_colon)
                    return

        try:
            # 保存Excel文件
            excel_file_path = os.path.join(WORK_FOLDER, "poc_list.xlsx")
            wb.save(excel_file_path)

        except PermissionError as e:
            print(f"Unexpected error: {e}")
            messagebox.showerror("错误", "请先关闭 'poc_list.xlsx' 文件！")
            return 1
        
        # 检查是否存在重复的id
        duplicate_ids = check_duplicates_in_excel(excel_file_path)
        print(duplicate_ids)
        
        if error_id:
            if error_name:
                if error_severity:
                    ERROR_MESSAGE = "以下POC文件 的 'id' 字段为空或不存在：\n" + "\n".join(error_id) + "\n\n以下POC文件的 'name' 字段为空或不存在：\n" + "\n".join(error_name) + "\n\n以下POC文件的 'severity' 字段为空或不存在：\n" + "\n".join(error_severity) + "\n\n请参考 'poc_list.xlsx' 文件并修改POC！"

                else:
                    ERROR_MESSAGE = "以下POC文件 的 'id' 字段为空或不存在：\n" + "\n".join(error_id) + "\n\n以下POC文件的 'name' 字段为空或不存在：\n" + "\n".join(error_name) + "\n\n请参考 'poc_list.xlsx' 文件并修改POC！"
            
            else:
                if error_severity:
                    ERROR_MESSAGE = "以下POC文件 的 'id' 字段为空或不存在：\n" + "\n".join(error_id) +  "\n\n以下POC文件的 'severity' 字段为空或不存在：\n" + "\n".join(error_severity) + "\n\n请参考 'poc_list.xlsx' 文件并修改POC！"

                else:
                    ERROR_MESSAGE = "以下POC文件 的 'id' 字段为空或不存在：\n" + "\n".join(error_id) + "\n\n请参考 'poc_list.xlsx' 文件并修改POC！"

        else:
            if error_name:
                if error_severity:
                    ERROR_MESSAGE = "以下POC文件的 'name' 字段为空或不存在：\n" + "\n".join(error_name) + "\n\n以下POC文件的 'severity' 字段为空或不存在：\n" + "\n".join(error_severity) + "\n\n请参考 'poc_list.xlsx' 文件并修改POC！"

                else:
                    ERROR_MESSAGE = "以下POC文件的 'name' 字段为空或不存在：\n" + "\n".join(error_name) + "\n\n请参考 'poc_list.xlsx' 文件并修改POC！"
            
            else:
                if error_severity:
                    ERROR_MESSAGE = "以下POC文件的 'severity' 字段为空或不存在：\n" + "\n".join(error_severity) + "\n\n请参考 'poc_list.xlsx' 文件并修改POC！"

                else:
                    if duplicate_ids:
                        print("以下 'id' 字段发现对应多个POC文件：")
                        print(duplicate_ids)
                        ERROR_MESSAGE = "以下 'id' 字段中发现多个POC文件：\n" + "\n".join(duplicate_ids) + "\n\n请参考 'poc_list.xlsx' 文件并修改POC！"

                    else:
                        CURRENT_POCLIST_RESULT = True
            
        wb.close()

    else:
        print("poc文件夹不存在。")
        messagebox.showerror("错误", "'poc' 文件夹不存在！")

# 检查Excel文件中是否存在重复的ID
def check_duplicates_in_excel(excel_file_path):
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb.active
    # 存储已经检查过的ID
    checked_ids = []
    # 存储已经发现的重复项
    duplicate_ids = []

    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        id_value = row[0]

        if id_value in checked_ids and id_value not in duplicate_ids:
            duplicate_ids.append(id_value)

        else:
            checked_ids.append(id_value)

    wb.close()
    return duplicate_ids

# 打开URL的函数
def open_url_left(event):
    # 确定点击的是哪一行
    selected_items = table.selection()

    # 检查是否按住了 Ctrl 或 Shift 键，如果是则直接返回，不执行操作
    if event.state & 4 or event.state & 1:
        return

    if event.y > 24:
        for item in selected_items:
            url = table.item(item, "values")[2]
            webbrowser.open(url)

def open_url_right(event):
    selected_items = table.selection()

    for item in selected_items:
        url = table.item(item, "values")[2]
        webbrowser.open(url)

# 复制URL的函数
def copy_url(event):
    selected_items = table.selection()
    urls = [table.item(item, "values")[2] for item in selected_items]
    root.clipboard_clear()
    root.clipboard_append("\n".join(urls))
    root.update()

# 显示右键菜单的函数
def show_right_click_menu(event):
    selected_items = table.selection()
    # 获取右键点击的项
    select_item = table.identify_row(event.y)

    if event.y > 24:
        if select_item not in selected_items:
            # 清除之前所有选中项
            for item in selected_items:
                table.selection_remove(item)

            # 选中右键点击的行
            table.selection_add(select_item)
            # 显示右键菜单
            right_click_menu.post(event.x_root, event.y_root)
            
        else:
            right_click_menu.post(event.x_root, event.y_root)

# 选择URL按钮删除线
def url_line(event):
    global URL_FILE_PATH

    url_content = target_url_entry.get().strip()

    if url_content and URL_FILE_PATH:
        url_delete_canvas.place(relx=0.249, rely=0.206, relwidth=0.46)

    else:
        url_delete_canvas.place(relx=1, rely=1, relwidth=0)

# 配置云平台
def set_cloud_platform():
    global API_KEY
    global CLOUD_UPLOAD_STATUS
    
    def save_api_key_config():
        global API_KEY

        API_KEY = api_key_entry.get()
        config.set("Settings", "API_Key", API_KEY)
        save_config()
    
    def save_cloud_upload_config():
        global CLOUD_UPLOAD_STATUS

        CLOUD_UPLOAD_STATUS = cloud_upload_button_var.get()
        config.set("Settings", "cloud_upload", str(CLOUD_UPLOAD_STATUS))  # 将布尔值转换为字符串保存
        save_config()

    # 创建一个Toplevel窗口，即弹窗
    cloud_window = tk.Toplevel(root)
    cloud_window.title("配置云平台")
    cloud_window.resizable(False, False)
    x = root.winfo_rootx() + 400
    y = root.winfo_rooty() + 80
    cloud_window.geometry(f"600x280+{x}+{y}")

    def open_sign_up(event):
        webbrowser.open("https://cloud.projectdiscovery.io/sign-up")
    
    def on_enter(event):
        api_key_label_2.config(foreground="blue")

    def on_leave(event):
        api_key_label_2.config(foreground="black")

    # 在弹窗中添加Label和Entry，用于输入API_Key
    api_key_label_1 = ttk.Label(cloud_window, font=("黑体", 18), text="请输入云平台API Key:")
    api_key_label_1.place(relx=0.04, rely=0.1, relwidth=0.89)
    api_key_label_2 = ttk.Label(cloud_window, font=("黑体", 18, "underline"), text="点击注册", cursor="hand2")
    api_key_label_2.place(relx=0.76, rely=0.1, relwidth=0.2)
    api_key_label_2.bind("<Button-1>", open_sign_up)
    api_key_label_2.bind("<Enter>", on_enter)
    api_key_label_2.bind("<Leave>", on_leave)
    api_key_entry = tk.Entry(cloud_window, font=("黑体", 20))
    api_key_entry.insert(0, API_KEY)
    api_key_entry.place(relx=0.044, rely=0.3, relwidth=0.89)

    def cloud_upload():
        global CLOUD_UPLOAD_STATUS

        CLOUD_UPLOAD_STATUS = cloud_upload_button_var.get()

    cloud_upload_button_style = ttk.Style()
    cloud_upload_button_style.configure("Custom.TCheckbutton", font=("黑体", 14, ""))
    cloud_upload_button_var = tk.BooleanVar()
    cloud_upload_button_var.set(CLOUD_UPLOAD_STATUS)
    cloud_upload_button = ttk.Checkbutton(cloud_window, text="扫描完成后将扫描结果上传到云平台", variable=cloud_upload_button_var, command=cloud_upload, style="Custom.TCheckbutton")
    cloud_upload_button.place(relx=0.04, rely=0.42, relwidth=0.89, relheight=0.2)

    # 在弹窗中添加"确定"按钮，点击后获取Entry中的值
    def ok_button_clicked():
        global API_KEY
        global CLOUD_UPLOAD_STATUS

        save_api_key_config()
        save_cloud_upload_config()

        if API_KEY and CLOUD_UPLOAD_STATUS:
            validate_api_key()

        cloud_window.destroy()
    
    def cancel_button_clicked():
        cloud_window.destroy()

    ok_button = ttk.Button(cloud_window, text="确定", command=ok_button_clicked, style="Custom.TButton")
    ok_button.place(relx=0.04, rely=0.64, relwidth=0.24, relheight=0.22)
    cancel_button = ttk.Button(cloud_window, text="取消", command=cancel_button_clicked, style="Custom.TButton")
    cancel_button.place(relx=0.7, rely=0.64, relwidth=0.24, relheight=0.22)
    # 设置焦点并阻塞主窗口
    cloud_window.focus_force()
    cloud_window.grab_set()
    root.wait_window(cloud_window)
    # 释放焦点
    cloud_window.grab_release()

# 检查云平台API_Key
def validate_api_key():
    global API_KEY
    global NUCLEI_PATH
    global WORK_FOLDER
    global ERROR_MESSAGE
    
    if not os.path.exists(NUCLEI_PATH):
        print("不存在nuclei.exe文件")
        return
    
    vbs_file_path = f"{WORK_FOLDER}/validate_API_Key.vbs"
    API_KEY= config.get("Settings", "API_Key")
    messagebox.showinfo("提示", "即将配置API Key，请等待cmd窗口自动关闭后再进行操作！")
    
    try:
        if not os.path.exists(vbs_file_path):
            vbs_code = """
            Set objShell = CreateObject("WScript.Shell")
            objShell.Run "cmd /k nuclei.exe -auth", 1, False
            WScript.Sleep 2000
            objShell.SendKeys "%s{ENTER}"
            objShell.SendKeys "exit{ENTER}"
            Set objShell = Nothing
            """ % API_KEY

            # 将VBScript代码写入到.vbs文件中
            with open(vbs_file_path, "w", encoding="utf-8") as file:
                file.write(vbs_code)

        # 使用subprocess执行VBScript文件
        subprocess.Popen(["cscript", vbs_file_path], shell=True)

    except subprocess.CalledProcessError as e:
        print(f"Unexpected error: {e}")

    except Exception as e:
        print(f"Unexpected error: {e}")
        show_nuclei_version("", "black")
        show_nuclei_status("Nuclei加载失败！", "red")
    
# 访问云平台
def open_cloud_platform():
    webbrowser.open("https://cloud.projectdiscovery.io/")

# 创建样式对象
button_style = ttk.Style()
# 定义新的样式，包括字体设置
button_style.configure("Custom.TButton", font=("黑体", 18, ""))

# 设置UA头按钮
set_ua_button = ttk.Button(root, text="设置UA头", takefocus=False, command=set_ua, style="Custom.TButton")
set_ua_button.place(relx=0.015, rely=0.01, relwidth=0.1, relheight=0.05)

# 设置代理按钮
set_proxy_button = ttk.Button(root, text="设置代理", takefocus=False, command=set_proxy, style="Custom.TButton")
set_proxy_button.place(relx=0.14, rely=0.01, relwidth=0.1, relheight=0.05)

# 显示代理状态标签
proxy_status_label = ttk.Label(root, font=("黑体", 20, "bold"))
proxy_status_label.place(relx=0.258, rely=0.01, relwidth=0.44, relheight=0.05)

# 新建POC按钮
new_poc_button = ttk.Button(root, text="新建POC", takefocus=False, command=new_poc, style="Custom.TButton")
new_poc_button.place(relx=0.015, rely=0.066, relwidth=0.1, relheight=0.05)

# 选择POC按钮
select_poc_button = ttk.Button(root, text="选择POC", takefocus=False, command=select_poc, style="Custom.TButton")
select_poc_button.place(relx=0.14, rely=0.066, relwidth=0.1, relheight=0.05)

# 选择POC状态
select_poc_entry = ttk.Label(root, font=("黑体", 20, "bold"))
select_poc_entry.place(relx=0.258, rely=0.066, relwidth=0.44, relheight=0.05)

# 选择POC结果标签
selectpoc_status_label = ttk.Label(root, text="", font=("黑体", 20, "bold"))
selectpoc_status_label.place(relx=0.258, rely=0.066, relwidth=0.44, relheight=0.05)

# 编辑POC按钮
target_poc_button = ttk.Button(root, text="编辑POC", takefocus=False, command=view_poc, style="Custom.TButton")
target_poc_button.place(relx=0.015, rely=0.122, relwidth=0.1, relheight=0.05)

# 查看POC列表按钮
view_pocs_button = ttk.Button(root, text="POC列表", takefocus=False, command=view_pocs, style="Custom.TButton")
view_pocs_button.place(relx=0.14, rely=0.122, relwidth=0.1, relheight=0.05)

# 检查POC结果标签
validation_status_label = ttk.Label(root, text="", font=("黑体", 20, "bold"))
validation_status_label.place(relx=0.258, rely=0.122, relwidth=0.5, relheight=0.05)

# 编辑URL按钮
target_url_button = ttk.Button(root, text="编辑URL", takefocus=False, command=view_urls, style="Custom.TButton")
target_url_button.place(relx=0.015, rely=0.178, relwidth=0.1, relheight=0.05)

# 选择URL按钮
select_file_button = ttk.Button(root, text="选择URL", takefocus=False, command=select_url_file, style="Custom.TButton")
select_file_button.place(relx=0.14, rely=0.178, relwidth=0.1, relheight=0.05)

# 输入框显示当前URL文件名
url_file_entry = ttk.Entry(root, width=50, font=("黑体", 20, ""), state="disabled")
url_file_entry.place(relx=0.258, rely=0.184, relwidth=0.44, relheight=0.04)

# 显示编辑状态
urls_status_label = ttk.Label(root, font=("黑体", 20, "bold"))
urls_status_label.place(relx=0.26, rely=0.188, relwidth=0.43, relheight=0.03)

# 配置云平台按钮
set_cloud_platform_button = ttk.Button(root, text="配置云平台", takefocus=False, command=set_cloud_platform, style="Custom.TButton")
set_cloud_platform_button.place(relx=0.015, rely=0.234, relwidth=0.1, relheight=0.05)

# 目标URL标签和输入框
target_url_label = ttk.Label(root, text="目标URL：", font=("黑体", 18, ""))
target_url_label.place(relx=0.162, rely=0.234, relwidth=0.1, relheight=0.05)
target_url_entry = tk.Entry(root, width=50, font=("黑体", 20, ""))
target_url_entry.place(relx=0.258, rely=0.244, relwidth=0.44, relheight=0.03)

# 创建Canvas
url_delete_canvas = tk.Canvas(root, height=2, highlightthickness=0, bg="black")
url_line(None)

# 添加横线切换的事件
target_url_entry.bind("<KeyRelease>", url_line)

# 开始扫描按钮
scan_button = ttk.Button(root, text="开始扫描", takefocus=False, command=lambda: Scan(target_url_entry.get()), style="Custom.TButton")
scan_button.place(relx=0.015, rely=0.29, relwidth=0.1, relheight=0.05)

# 终止扫描按钮
stop_scan_button = ttk.Button(root, text="终止扫描", takefocus=False, command=stop_scan, style="Custom.TButton")
stop_scan_button.place(relx=0.14, rely=0.29, relwidth=0.1, relheight=0.05)

# 扫描状态
scan_status_label = ttk.Label(root, text="未开始扫描", font=("黑体", 20, "bold"))
scan_status_label.place(relx=0.258, rely=0.29, relwidth=0.44, relheight=0.05)

# 访问云平台按钮
open_cloud_platform_button = ttk.Button(root, text="访问云平台", takefocus=False, command=open_cloud_platform, style="Custom.TButton")
open_cloud_platform_button.place(relx=0.72, rely=0.29, relwidth=0.11, relheight=0.05)

# 查看扫描结果按钮
results_button = ttk.Button(root, text="查看扫描结果", takefocus=False, command=view_results, style="Custom.TButton")
results_button.place(relx=0.856, rely=0.29, relwidth=0.12, relheight=0.05)

# 检查Nuclei状态
nuclei_status_label = ttk.Label(root, text="", font=("黑体", 20, "bold"))
nuclei_status_label.place(relx=0.72, rely=0.186, relwidth=0.44, relheight=0.04)

# 检查Nuclei版本
nuclei_version_label = ttk.Label(root, text="", font=("黑体", 20, "bold"))
nuclei_version_label.place(relx=0.72, rely=0.238, relwidth=0.44, relheight=0.04)

nuclei_logo_label = ttk.Label(root, text="", font=("黑体", 20))
nuclei_logo_label.place(relx=0.72, rely=0, relwidth=0.5, relheight=0.14)


# 扫描结果表格
columns = ("Name", "Severity", "URL")
table = ttk.Treeview(root, columns=columns, show="headings", height=18, style="Custom.Treeview")
column_style = ttk.Style()
column_style.configure("Custom.Treeview", rowheight=38, font=("黑体", 18))

def resize_treeview_columns(treeview):
    head_style = ttk.Style()
    head_style.configure("Treeview.Heading", font=("黑体", 19, "bold"))

resize_treeview_columns(table)

def table_color(): # 表格栏隔行显示不同颜色函数
    items = table.get_children()
    for index, iid in enumerate(items):
        tag = "even" if index % 2 == 0 else ""
        table.item(iid, tag=tag)

def fixed_map(option):
    return [elm for elm in column_color_style.map("Treeview", query_opt=option) if elm[:2] != ("!disabled", "!selected")]

column_color_style = Style()
column_color_style.map("Treeview", foreground=fixed_map("foreground"), background=fixed_map("background"))
light_gray = "#F5F5F5"
table.tag_configure("even", background=light_gray)

# 设置表头
for col in columns:
    table.heading(col, text=col, anchor="center", command=lambda c=col: sort_treeview(table, c))
    table.bind("<B1-Motion>", lambda e: "break")

# 设置表格布局
table.column("Name", width=620)
table.column("Severity", width=140)
table.column("URL", width=775)
table.place(relx=0.015, rely=0.366)

# 在表格内添加边框
table_canvas_x = tk.Canvas(root, height=1, highlightthickness=0, bg="gray")
table_canvas_x.place(relx=0.015, rely=0.387, relwidth=0.96)
table_canvas_y_1 = tk.Canvas(root, width=1, highlightthickness=0, bg="gray")
table_canvas_y_1.place(relx=0.4022, rely=0.366, relheight=0.592)
table_canvas_y_2 = tk.Canvas(root, width=1, highlightthickness=0, bg="gray")
table_canvas_y_2.place(relx=0.4903, rely=0.366, relheight=0.592)

# 列的排序状态
sort_states = {col: False for col in columns}

# 为表格添加双击事件
table.bind("<Double-1>", open_url_left)

# 创建右键菜单
right_click_menu = tk.Menu(root, tearoff=0)
right_click_menu.config(font=("黑体", 13))
right_click_menu.add_command(label="打开URL", command=lambda event=None: open_url_right(event))
right_click_menu.add_command(label="复制URL", command=lambda event=None: copy_url(event))

# 绑定右键点击事件
table.bind("<Button-3>", show_right_click_menu)

# 创建垂直滚动条
y_scrollbar = ttk.Scrollbar(root, orient="vertical", command=table.yview)
y_scrollbar.place(relx=0.975, rely=0.37, relheight=0.603)
cover_label = ttk.Label(root, text="")
cover_label.place(relx=0.974, rely=0.958, relheight=0.02)

# 配置表格与垂直滚动条的关联
table.configure(yscrollcommand=y_scrollbar.set)
y_scrollbar.lower()

update_proxy_status()
show_nuclei_logo("                     __     _\n   ____  __  _______/ /__  (_)\n  / __ \/ / / / ___/ / _ \/ /\n / / / / /_/ / /__/ /  __/ /\n/_/ /_/\__,_/\___/_/\___/_/   ","black")
show_selectpoc_status("当前未选择POC！", "red")
Load()
root.mainloop()
